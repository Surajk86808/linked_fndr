# data/xlsx_writer.py  -  XLSX output with formatting, hyperlinks, and summary sheet

import os
from collections import Counter

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import config
from core.scraper import Lead
from utils.logger import log

HEADERS = [
    "Name", "Current Title", "Company", "Location", "LinkedIn URL",
    "Website", "Has Website", "Connections", "Headline",
    "Role Start Date", "Days Since Role Start", "Role Recency",
    "About (Preview)", "Skills", "Education",
    "Score (0-100)", "Score Breakdown", "Priority",
    "Keyword", "Country", "Scraped Date",
]

COLUMN_WIDTHS = [
    22, 28, 25, 20, 45,
    40, 12, 15, 40,
    18, 20, 15,
    50, 40, 30,
    12, 60, 10,
    20, 18, 14,
]

PRIORITY_COLORS = {
    "HIGH":   "00C851",
    "MEDIUM": "FFD700",
    "LOW":    "FF4444",
}

HEADER_FILL  = PatternFill("solid", fgColor="1B3A6B")
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=10)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _thin_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def init_xlsx(path: str = config.OUTPUT_XLSX):
    """Create the XLSX file with formatted headers if it doesn't exist."""
    if os.path.exists(path):
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    for col_idx, (header, width) in enumerate(zip(HEADERS, COLUMN_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border    = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    wb.save(path)
    log(f"Output XLSX created: {path}", "OK")


def save_lead(lead: Lead, country: str = "", path: str = config.OUTPUT_XLSX):
    """Append a single Lead row to the XLSX with priority colour coding."""
    if not os.path.exists(path):
        init_xlsx(path)

    wb = load_workbook(path)
    ws = wb.active
    next_row = ws.max_row + 1

    row_data = [
        lead.name, lead.current_title, lead.company, lead.location,
        lead.linkedin_url, lead.website, lead.has_website, lead.connections,
        lead.headline, lead.role_start_date, lead.role_start_days,
        lead.role_recency_label, lead.about, lead.skills, lead.education,
        lead.score, lead.score_breakdown, lead.priority,
        lead.keyword_matched, country, lead.scraped_date,
    ]

    priority_color = PRIORITY_COLORS.get(lead.priority, "FFFFFF")
    priority_fill  = PatternFill("solid", fgColor=priority_color)

    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=next_row, column=col_idx, value=value)
        cell.border    = _thin_border()
        cell.alignment = Alignment(vertical="top", wrap_text=False)

        header = HEADERS[col_idx - 1]

        if header == "Priority":
            cell.fill = priority_fill
            font_color = "FFFFFF" if lead.priority != "MEDIUM" else "000000"
            cell.font  = Font(bold=True, color=font_color)

        if header == "LinkedIn URL" and value:
            cell.hyperlink = value
            cell.font = Font(color="0563C1", underline="single")

        if header == "Website" and value:
            cell.hyperlink = value
            cell.font = Font(color="0563C1", underline="single")

    wb.save(path)


def add_summary_sheet(path: str = config.OUTPUT_XLSX):
    """Add a Summary sheet with aggregate stats to the final XLSX."""
    if not os.path.exists(path):
        return

    wb = load_workbook(path)

    if "Summary" in wb.sheetnames:
        del wb["Summary"]

    ws_leads = wb["Leads"]
    ws_sum   = wb.create_sheet("Summary", 0)

    rows = list(ws_leads.iter_rows(min_row=2, values_only=True))
    if not rows:
        wb.save(path)
        return

    priority_idx = HEADERS.index("Priority")
    country_idx  = HEADERS.index("Country")

    priorities = Counter(r[priority_idx] for r in rows if r[priority_idx])
    countries  = Counter(r[country_idx]  for r in rows if r[country_idx])

    # Title
    ws_sum["A1"] = "NexviaTech Lead Scrape Summary"
    ws_sum["A1"].font = Font(bold=True, size=14, color="1B3A6B")

    # Totals
    ws_sum["A3"] = "Total Leads"
    ws_sum["B3"] = len(rows)
    ws_sum["A3"].font = Font(bold=True)

    ws_sum["A4"] = "HIGH Priority"
    ws_sum["B4"] = priorities.get("HIGH", 0)
    ws_sum["A4"].font = Font(color="00C851", bold=True)

    ws_sum["A5"] = "MEDIUM Priority"
    ws_sum["B5"] = priorities.get("MEDIUM", 0)
    ws_sum["A5"].font = Font(color="CC9900", bold=True)

    ws_sum["A6"] = "LOW Priority"
    ws_sum["B6"] = priorities.get("LOW", 0)
    ws_sum["A6"].font = Font(color="FF4444", bold=True)

    # Country breakdown
    ws_sum["A8"] = "Leads by Country"
    ws_sum["A8"].font = Font(bold=True, size=11)
    for i, (country, count) in enumerate(countries.most_common(), start=9):
        ws_sum.cell(row=i, column=1, value=country)
        ws_sum.cell(row=i, column=2, value=count)

    ws_sum.column_dimensions["A"].width = 25
    ws_sum.column_dimensions["B"].width = 12

    wb.save(path)
    log("Summary sheet added to XLSX.", "OK")
